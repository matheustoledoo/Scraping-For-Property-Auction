# ============================================================
# 1. Importação de módulos e configuração inicial
# ============================================================

from selenium import webdriver  # Para controlar o navegador
from selenium.webdriver.common.by import By  # Para localizar elementos
from selenium.webdriver.chrome.service import Service  # Para configurar o ChromeDriver
from selenium.webdriver.chrome.options import Options  # Para definir opções do Chrome (ex: modo headless)
import time  # Para utilizar time.sleep()
import re  # Para limpeza de caracteres indesejados
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # Para remover caracteres ilegais no Excel

def print_header(message):
    print("\n" + "=" * 70)
    print(f"{message}".center(70))
    print("=" * 70 + "\n")

def clean_text(text):
    if text is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", text)


# ============================================================
# 2. Configuração do Chrome (Headless)
# ============================================================

print_header("Chrome Configuration (Headless) para Mega Leilões")
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

chrome_driver_path = "C:\\Users\\mathe\\Desktop\\chromedriver-win64\\chromedriver.exe"  # Atualize para o caminho do seu chromedriver
service = Service(executable_path=chrome_driver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

# ============================================================
# 3. Coleta dos Links dos Imóveis das Páginas Desejadas
# ============================================================

# URL base para Mega Leilões – observe que o parâmetro de página é usado
base_url = "https://www.megaleiloes.com.br/imoveis?tov=igbr&valor_max=5000000&tipo%5B0%5D=1&pagina={page}"

print_header("Coletando Links dos Imóveis - Mega Leilões")
paginas_input = input("Digite o número de páginas a serem raspadas (ou 'todas'): ")
if paginas_input.lower() == "todas":
    total_pages = None
else:
    total_pages = int(paginas_input)
print(f"[INFO] Páginas a serem raspadas: {'Todas' if total_pages is None else total_pages}")

all_links = []  # Lista para armazenar os links dos imóveis
status_dict = {}  # Dicionário para armazenar o status de cada imóvel
current_page = 1

while True:
    if total_pages is not None and current_page > total_pages:
        break

    current_url = base_url.format(page=current_page)
    print_header(f"Coletando Links - Página {current_page}")
    print(f"[INFO] Acessando: {current_url}")
    driver.get(current_url)
    time.sleep(4)  # Ajuste conforme necessário

    # Supondo que cada imóvel esteja contido em uma div com as classes "col-sm-6 col-md-4 col-lg-3"
    imovel_cards = driver.find_elements(By.XPATH, '//div[contains(@class, "col-sm-6 col-md-4 col-lg-3")]')
    print(f"[INFO] Itens encontrados na página {current_page}: {len(imovel_cards)}")

    if len(imovel_cards) == 0:
        print("[INFO] Nenhum item encontrado. Encerrando coleta.")
        break

    # Itera sobre os cartões (cards)
    for index, card in enumerate(imovel_cards, start=1):
        try:
            # Extrai o status (por exemplo, um <span> com classe "card-status")
            status_element = card.find_element(By.XPATH, './/div[contains(@class, "card-status")]')
            status_text = status_element.text.strip()
        except Exception as e:
            status_text = ""

        try:
            # Extrai o link para o imóvel (por exemplo, de um <a> com classe "card-title")
            link_element = card.find_element(By.XPATH, './/a[contains(@class, "card-title")]')
            link = link_element.get_attribute('href')
            all_links.append(link)
            status_dict[link] = status_text  # Armazena o status
            print(f"  [OK] Card {index}: {link} (Status: {status_text})")
        except Exception as e:
            print(f"  [ERRO] Card {index}: {e}")

    print(f"[INFO] Total de links coletados até a Página {current_page}: {len(all_links)}")
    current_page += 1

print_header("Total de Links Coletados")
print(f"[INFO] Total de imóveis coletados: {len(all_links)}")

# ============================================================
# 4. Processamento dos Imóveis (Extração dos Dados)
# ============================================================

all_imoveis_data = []  # Lista que conterá os dados dos imóveis

for i, link in enumerate(all_links, start=1):
    print_header(f"Processando Imóvel {i}/{len(all_links)}")
    print(f"[INFO] URL: {link}")
    # Abre o link em uma nova aba
    driver.execute_script("window.open(arguments[0]);", link)
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(3)  # Aguarda o carregamento

    try:
        # Título do imóvel
        titulo_leilao = driver.find_element(By.XPATH, '//h1[contains(@class, "section-header")]').text.strip()
        print(f"  [OK] Título: {titulo_leilao}")
    except Exception as e:
        titulo_leilao = None
        print(f"  [ERRO] Título: {e}")

    try:
        # Tipo de leilão
        tipo_leilao = driver.find_element(By.XPATH, '//div[contains(@class, "batch-type")]').text.strip()
        print(f"  [OK] Tipo: {tipo_leilao}")
    except Exception as e:
        tipo_leilao = None
        print(f"  [ERRO] Tipo: {e}")

    try:
        # Número do processo
        numero_processo = driver.find_element(
            By.XPATH,
            '/html/body/div[3]/div[3]/div[2]/div[2]/div/div/div[2]/div[1]/div[2]/a'
        ).text.strip()
        print(f"  [OK] Nº Processo: {numero_processo}")
    except Exception as e:
        numero_processo = None
        print(f"  [ERRO] Nº Processo: {e}")

    try:
        # Valor do imóvel
        valor_imovel = driver.find_element(By.XPATH, '//div[contains(@class, "value")]').text.strip()
        print(f"  [OK] Valor: {valor_imovel}")
    except Exception as e:
        valor_imovel = None
        print(f"  [ERRO] Valor: {e}")

    try:
        # Extração do link para o Edital usando XPath absoluto
        edital_leilao = driver.find_element(
            By.XPATH, '/html/body/div[3]/div[3]/div[3]/div[3]/div[2]/a[2]'
        ).get_attribute('href')
        print(f"  [OK] Edital: {edital_leilao}")
    except Exception as e:
        edital_leilao = None
        print(f"  [ERRO] Edital: {e}")

    try:
        # Extração do link do Laudo de Avaliação
        laudo_avaliacao = driver.find_element(
            By.XPATH, '/html/body/div[3]/div[3]/div[3]/div[3]/div[2]/a[3]'
        ).get_attribute('href')
        print(f"  [OK] Laudo de Avaliação: {laudo_avaliacao}")
    except Exception as e:
        laudo_avaliacao = None
        print(f"  [ERRO] Laudo de Avaliação: {e}")

    try:
        # Extração do link da Matrícula
        matricula = driver.find_element(
            By.XPATH, '/html/body/div[3]/div[3]/div[3]/div[3]/div[2]/a[4]'
        ).get_attribute('href')
        print(f"  [OK] Matrícula: {matricula}")
    except Exception as e:
        matricula = None
        print(f"  [ERRO] Matrícula: {e}")

    # Monta o dicionário de documentos
    documentos = {
        "Edital": edital_leilao,
        "Laudo de Avaliação": laudo_avaliacao,
        "Matricula": matricula
    }

    try:
        # Extração da descrição do imóvel
        descricao_lote = driver.find_element(By.XPATH, '//div[contains(@class, "description")]').text.strip()
        print("  [OK] Descrição do Imóvel extraída.")
    except Exception as e:
        descricao_lote = None
        print(f"  [ERRO] Descrição do Imóvel: {e}")

    # Recupera o status previamente armazenado para este link
    status_leilao = status_dict.get(link, "")
    imovel_info = {
        "link": link,
        "titulo_leilao": titulo_leilao,
        "tipo_leilao": tipo_leilao,
        "numero_processo": numero_processo,
        "valor_imovel": valor_imovel,
        "edital_leilao": edital_leilao,
        "documentos": documentos,
        "descricao_lote": descricao_lote,
        "status": status_leilao
    }
    all_imoveis_data.append(imovel_info)

    print_header(f"Dados Extraídos do Imóvel {i}")
    for chave, valor in imovel_info.items():
        print(f"{chave:20}: {valor}")

    # Fecha a aba do imóvel e retorna à aba principal
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(1)

# ============================================================
# 5. Exportação para Excel (XLSX)
# ============================================================

import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

print_header("Preparando os dados para exportação para XLSX")

# Lista de documentos padrão para as colunas do Excel – agora com os 3 campos desejados
docs_padrao = [
    "Edital",
    "Laudo de Avaliação",
    "Matricula"
]

# Converte os dados coletados para um DataFrame
dados_formatados = []
for idx, item in enumerate(all_imoveis_data, start=1):
    documentos = item.get("documentos") or {}
    row = {
        "ID": f"Imóvel {idx}",
        "Título do Leilão": clean_text(str(item.get("titulo_leilao", ""))),
        "Tipo de Leilão": clean_text(str(item.get("tipo_leilao", ""))),
        "Número do Processo": clean_text(str(item.get("numero_processo", ""))),
        "Valor do Imóvel": clean_text(str(item.get("valor_imovel", ""))),
        "Link do Edital": clean_text(str(item.get("edital_leilao", ""))),
        "Link do Imóvel": clean_text(str(item.get("link", ""))),
        "Descrição do Imóvel": clean_text(str(item.get("descricao_lote", ""))),
        "Status": clean_text(str(item.get("status", "")))
    }
    for doc in docs_padrao:
        row[doc] = clean_text(str(documentos.get(doc, "")))
    dados_formatados.append(row)

# Define a ordem das colunas
colunas = [
    "ID", "Título do Leilão", "Tipo de Leilão", "Número do Processo",
    "Valor do Imóvel", "Link do Edital", "Link do Imóvel", "Descrição do Imóvel", "Status"
] + docs_padrao

df = pd.DataFrame(dados_formatados, columns=colunas)

# ============================================================
# 6. Escolha do local para salvar o arquivo XLSX
# ============================================================

print_header("Escolha onde salvar a planilha XLSX")
root = tk.Tk()
root.withdraw()
root.lift()
root.attributes("-topmost", True)

nome_padrao = "leiloes_megaleiloes_formatado.xlsx"
caminho_arquivo = filedialog.asksaveasfilename(
    initialfile=nome_padrao,
    defaultextension=".xlsx",
    filetypes=[("Planilhas Excel", "*.xlsx")],
    title="Salvar planilha como"
)

if caminho_arquivo:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leilões Mega"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        cell_font = Font(color="000000")
        cell_alignment = Alignment(wrap_text=True, vertical="top")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        link_columns = ["Link do Edital", "Link do Imóvel"] + docs_padrao

        # Escreve os dados na planilha
        for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                header_name = df.columns[c_idx - 1] if r_idx > 1 else None
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = cell_alignment

                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = cell_font
                    if header_name in link_columns and str(value).strip().lower().startswith("http"):
                        cell.hyperlink = value
                        cell.font = Font(color="0000FF", underline="single")

        # Ajuste de largura das colunas
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 5, 80)

        wb.save(caminho_arquivo)
        print(f"[SUCESSO] Planilha XLSX salva com sucesso em:\n{caminho_arquivo}")

    except Exception as e:
        print(f"[ERRO] Erro ao salvar a planilha: {e}")
else:
    print("[CANCELADO] Nenhum arquivo foi salvo.")

# ============================================================
# 7. Finaliza o Navegador
# ============================================================

print_header("Finalizando o Scraping e Fechando o Navegador...")
driver.quit()



