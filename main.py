from fastapi import FastAPI, Form, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
import io
import pandas as pd

# scrapers
from scrapers.megaleiloes import run as scrape_mega
from scrapers.vivaleiloes import run as scrape_viva
from scrapers.saraivaleiloes import run as scrape_saraiva
from scrapers.leilaobrasil import run as scrape_brasil
from scrapers.mgs import run as scrape_mgs

# Excel styling
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

app = FastAPI()

@app.get("/", response_class=FileResponse)
async def index():
    return FileResponse("frontend/index.html")

@app.post("/scrape")
async def scrape(
    sites:         list[str] = Form(...),
    pages:         str       = Form(...),
    min_valor:     str       = Form(None),
    max_valor:     str       = Form(None),
    leilao_types:  list[str] = Form(None),
    states:        list[str] = Form(None),
    cities:        list[str] = Form(None),      # ← adicionado
    include_summary: bool    = Form(False)
):
    # --- parâmetros ---
    pages_int = -1 if pages.lower() == "todas" else int(pages)
    min_v     = float(min_valor) if min_valor else None
    max_v     = float(max_valor) if max_valor else None

    # --- coleta todos os dados sem filtrar estados/cidades ---
    dfs: dict[str, pd.DataFrame] = {}
    if "mega_leiloes"  in sites:
        dfs["Mega Leilões"]  = scrape_mega(pages_int)
    if "viva_leiloes"   in sites:
        dfs["Viva Leilões"]  = scrape_viva(pages_int)
    if "saraiva_leiloes" in sites:
        dfs["Saraiva Leilões"] = scrape_saraiva(pages_int)
    if "brasil_leiloes" in sites:
        dfs["Brasil Leilões"] = scrape_brasil(pages_int)
    if "mgs_leiloes" in sites:
        dfs["MGS Leilões"] = scrape_mgs(pages_int)

    # --- prepara listas para filtro de estados e cidades ---
    selected_ufs    = [uf.upper() for uf in (states or [])]
    selected_cities = [c.upper()  for c  in (cities or [])]   # ← adicionado

    # --- aplica filtros em cada DataFrame ---
    for name, df in dfs.items():
        if df.empty:
            continue

        # converte valor_imovel para float
        df["_valor_num"] = (
            df["valor_imovel"]
              .str.replace(r"[^\d,]", "", regex=True)
              .str.replace(",", ".")
              .astype(float)
        )

        # valor mínimo/máximo
        if min_v is not None:
            df = df[df["_valor_num"] >= min_v]
        if max_v is not None:
            df = df[df["_valor_num"] <= max_v]

        # tipo de leilão
        if leilao_types:
            df = df[df["tipo_leilao"].isin(leilao_types)]

        # filtro de estados
        if selected_ufs:
            df = df[df["estado"].str.upper().isin(selected_ufs)]

        # ** filtro de cidades **
        if selected_cities:
            df = df[df["cidade"].str.upper().isin(selected_cities)]

        dfs[name] = df

    # se não restou nada, retorna erro
    if not any(not df.empty for df in dfs.values()):
        raise HTTPException(400, "Nenhum imóvel encontrado com esses filtros.")

    # --- gera Excel ---
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            # grava sem coluna auxiliar
            df.drop(columns=["_valor_num"], errors="ignore") \
              .to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            max_col, max_row = ws.max_column, ws.max_row

            # tabela estilizada
            tbl = Table(
                displayName=sheet_name.replace(" ", "") + "_Tbl",
                ref=f"A1:{chr(64+max_col)}{max_row}"
            )
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tbl)

            # hyperlinks
            link_cols = {"link", "edital_leilao", "laudo_avaliacao", "matricula"}
            for idx, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=max_row), start=1):
                hdr = ws.cell(row=1, column=idx).value
                if hdr in link_cols:
                    for cell in col_cells:
                        if isinstance(cell.value, str) and cell.value.startswith("http"):
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0000FF", underline="single")

            # formatação condicional (valores acima da média)
            if max_row >= 2 and "_valor_num" in df.columns:
                avg = df["_valor_num"].mean()
                ci  = df.columns.get_loc("_valor_num") + 1
                rng = f"{chr(64+ci)}2:{chr(64+ci)}{max_row}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator="greaterThan",
                        formula=[str(avg)],
                        fill=PatternFill("solid", fgColor="FFC7CE")
                    )
                )

            ws.freeze_panes = "A2"

            # ajusta largura de colunas
            for col in ws.columns:
                lengths = [len(str(c.value)) for c in col if c.value]
                width   = min(max(lengths + [0]) + 5, 30)
                ws.column_dimensions[col[0].column_letter].width = width

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="leiloes.xlsx"'}
    )

# uvicorn main:app --reload --host 127.0.0.1 --port 8080