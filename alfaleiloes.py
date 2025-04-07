# ============================================================
# 1. Importação de módulos e configuração inicial
# ============================================================

from selenium import webdriver  # Para controlar o navegador
from selenium.webdriver.common.by import By  # Para localizar elementos
from selenium.webdriver.chrome.service import Service  # Para configurar o ChromeDriver
from selenium.webdriver.chrome.options import Options  # Para definir opções do Chrome (ex: modo headless)
import time  # Para utilizar time.sleep()
import re  # Para limpeza de caracteres indesejados
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # Para remover caracteres ilegais no Excel


def print_header(message):
    print("\n" + "=" * 70)
    print(f"{message}".center(70))
    print("=" * 70 + "\n")


# Função para limpar textos removendo caracteres ilegais para o Excel
def clean_text(text):
    if text is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", text)


# ============================================================
# 2. Configuração do Chrome (Headless)
# ============================================================

print_header("Chrome Configuration (Headless)")
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

chrome_driver_path = "C:\\Users\\mathe\\Desktop\\chromedriver-win64\\chromedriver.exe"
service = Service(executable_path=chrome_driver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

# ============================================================
# 3. Coleta dos Links dos Imóveis de Todas as Páginas Desejadas
# ============================================================

base_url = "https://www.alfaleiloes.com/leiloes/?&page={page}&categoria=35&categoria=18&categoria=19&categoria=24&categoria=23&categoria=26&categoria=27&search="

print_header("Coletando Links dos Imóveis")
paginas_input = input("Digite o número de páginas a serem raspadas (ou 'todas'): ")
if paginas_input.lower() == "todas":
    total_pages = None
else:
    total_pages = int(paginas_input)
print(f"[INFO] Páginas a serem raspadas: {'Todas' if total_pages is None else total_pages}")

all_links = []
status_dict = {}  # Armazena o status de cada link
current_page = 1

while True:
    if total_pages is not None and current_page > total_pages:
        break

    current_url = base_url.format(page=current_page)
    print_header(f"Coletando Links - Página {current_page}")
    print(f"[INFO] Acessando: {current_url}")
    driver.get(current_url)
    time.sleep(5)

    # Itera sobre os cards de leilão
    leilao_items = driver.find_elements(By.XPATH, '//div[@class="cards-wrapper"]/div[@class="home-leiloes-cards"]')
    print(f"[INFO] Itens encontrados na página {current_page}: {len(leilao_items)}")

    if len(leilao_items) == 0:
        print("[INFO] Nenhum item encontrado. Encerrando coleta.")
        break

    for index, item in enumerate(leilao_items, start=1):
        try:
            # Extrai o status a partir do elemento "card-status" e seu <p> interno
            status_element = item.find_element(By.CLASS_NAME, "card-status")
            status_text = status_element.find_element(By.TAG_NAME, "p").text.strip()
        except Exception as e:
            status_text = ""
        # Se o status for "Vendido", ignora este item
        if status_text.lower() == "vendido":
            print(f"  [INFO] Item {index} com status 'Vendido'. Ignorando.")
            continue

        try:
            link_element = item.find_element(By.XPATH, './/a[@class="btn-card"]')
            link = link_element.get_attribute('href')
            all_links.append(link)
            status_dict[link] = status_text  # Armazena o status do leilão (Aberto ou Futuro)
            print(f"  [OK] Item {index}: {link} (Status: {status_text})")
        except Exception as e:
            print(f"  [ERRO] Item {index}: {e}")

    print(f"[INFO] Total de links coletados até a Página {current_page}: {len(all_links)}")
    current_page += 1

print_header("Total de Links Coletados")
print(f"[INFO] Total de imóveis coletados: {len(all_links)}")

# ============================================================
# 4. Processamento dos Imóveis (Extração dos Dados)
# ============================================================
# (Esta parte permanece inalterada, com acréscimo do status)
all_imoveis_data = []

for i, link in enumerate(all_links, start=1):
    print_header(f"Processando Imóvel {i}/{len(all_links)}")
    print(f"[INFO] URL: {link}")
    driver.execute_script("window.open(arguments[0]);", link)
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(3)

    try:
        titulo_leilao = driver.find_element(By.CLASS_NAME, "title-lote-leiloes").text
        print(f"  [OK] Título: {titulo_leilao}")
    except Exception as e:
        titulo_leilao = None
        print(f"  [ERRO] Título: {e}")

    try:
        tipo_leilao = driver.find_element(By.XPATH, '//*[@id="lotes"]/div[1]/div/h1').text
        print(f"  [OK] Tipo: {tipo_leilao}")
    except Exception as e:
        tipo_leilao = None
        print(f"  [ERRO] Tipo: {e}")

    try:
        numero_processo = driver.find_element(By.XPATH, '//*[@id="lotes"]/div[1]/div/div[4]/div[1]/a').text
        print(f"  [OK] Nº Processo (via a): {numero_processo}")
    except Exception as e:
        try:
            numero_processo = driver.find_element(By.XPATH, '//*[@id="lotes"]/div[1]/div/div[4]/div[1]/p[2]').text
            print(f"  [OK] Nº Processo (via p[2]): {numero_processo}")
        except Exception as e2:
            numero_processo = None
            print(f"  [ERRO] Nº Processo: {e2}")

    try:
        valor_imovel = driver.find_element(By.CLASS_NAME, "line-through").text
        print(f"  [OK] Valor (via classe): {valor_imovel}")
    except Exception as e:
        try:
            valor_imovel = driver.find_element(By.XPATH,
                                               '/html/body/div[2]/section[2]/div[1]/div/div[5]/ul/li[3]/p').text
            print(f"  [OK] Valor (via XPath): {valor_imovel}")
        except Exception as e2:
            valor_imovel = None
            print(f"  [ERRO] Valor: {e2}")

    try:
        edital_leilao = driver.find_element(By.XPATH,
                                            '//a[contains(translate(text(),"EDITAL","edital"), "edital")]').get_attribute('href')
        print(f"  [OK] Edital: {edital_leilao}")
    except Exception as e:
        edital_leilao = None
        print(f"  [ERRO] Edital: {e}")

    try:
        link_docs = driver.find_element(By.XPATH,
                                        '//a[contains(translate(text(),"DOCUMENTOS","documentos"), "documentos")]')
        link_docs.click()
        time.sleep(2)
        docs_container = driver.find_element(By.CLASS_NAME, "modal-body-doc")
        doc_links = docs_container.find_elements(By.TAG_NAME, "a")
        nomes_docs = [
            "Certidão de Matrícula",
            "Laudo de Avaliação",
            "Débitos Tributários",
            "Débito Exequendo/Condominial",
            "Manual de Participação"
        ]
        documentos_dict = {}
        for j, doc_link in enumerate(doc_links):
            href = doc_link.get_attribute('href')
            if j < len(nomes_docs):
                documentos_dict[nomes_docs[j]] = href
            else:
                documentos_dict[f"Documento {j + 1}"] = href
        print("  [OK] Documentos:")
        for nome, link_doc in documentos_dict.items():
            print(f"       {nome:30}: {link_doc}")
        documentos = documentos_dict
    except Exception as e:
        documentos = None
        print(f"  [ERRO] Documentos: {e}")

    try:
        descricao_lote = driver.find_element(By.CLASS_NAME, "content").text
        print("  [OK] Descrição do Lote extraída.")
    except Exception as e:
        descricao_lote = None
        print(f"  [ERRO] Descrição do Lote: {e}")

    # Recupera o status previamente armazenado para este link
    status_leilao = status_dict.get(link, "")
    imovel_info = {
        "link": link,
        "titulo_leilao": titulo_leilao,
        "tipo_leilao": tipo_leilao,
        "numero_processo": numero_processo,
        "valor_imovel": valor_imovel,
        "edital_leilao": edital_leilao,
        "documentos": documentos,
        "descricao_lote": descricao_lote,
        "status": status_leilao
    }
    all_imoveis_data.append(imovel_info)
    print_header(f"Dados Extraídos do Imóvel {i}")
    for chave, valor in imovel_info.items():
        print(f"{chave:20}: {valor}")

    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(1)

# ============================================================
# (Opcional: DEBUG) Verifica o conteúdo coletado
# ============================================================
print_header("DEBUG: Verificando o Conteúdo de all_imoveis_data")
for idx, imovel in enumerate(all_imoveis_data, start=1):
    print(f"\nImóvel #{idx}:")
    for k, v in imovel.items():
        print(f"  {k} => {v}")

# ============================================================
# 5. Exportação para Excel (XLSX estilizado e organizado)
# ============================================================
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

print_header("Preparando os dados para exportação")

# Lista de documentos para as colunas do Excel
docs_padrao = [
    "Certidão de Matrícula",
    "Laudo de Avaliação",
    "Débitos Tributários",
    "Débito Exequendo/Condominial",
    "Manual de Participação"
]

# Converte os dados coletados (all_imoveis_data) para um formato adequado ao DataFrame,
# mapeando as chaves originais para nomes de colunas “amigáveis” e aplicando clean_text.
dados_formatados = []
for idx, item in enumerate(all_imoveis_data, start=1):
    documentos = item.get("documentos") or {}
    row = {
        "ID": f"Imóvel {idx}",
        "Título do Leilão": clean_text(str(item.get("titulo_leilao", ""))),
        "Tipo de Leilão": clean_text(str(item.get("tipo_leilao", ""))),
        "Número do Processo": clean_text(str(item.get("numero_processo", ""))),
        "Valor do Imóvel": clean_text(str(item.get("valor_imovel", ""))),
        "Link do Edital": clean_text(str(item.get("edital_leilao", ""))),
        "Link do Imóvel": clean_text(str(item.get("link", ""))),
        "Descrição do Lote": clean_text(str(item.get("descricao_lote", ""))),
        "Status": clean_text(str(item.get("status", "")))
    }
    # Acrescenta as colunas de documentos
    for doc in docs_padrao:
        row[doc] = clean_text(str(documentos.get(doc, "")))
    dados_formatados.append(row)

# Define a ordem desejada das colunas (incluindo "Status")
colunas = [
    "ID", "Título do Leilão", "Tipo de Leilão", "Número do Processo",
    "Valor do Imóvel", "Link do Edital", "Link do Imóvel", "Descrição do Lote",
    "Status"
] + docs_padrao

df = pd.DataFrame(dados_formatados, columns=colunas)

# ============================================================
# 6. Escolha onde salvar o arquivo XLSX (janela de diálogo)
# ============================================================
print_header("Escolha onde salvar a planilha XLSX profissional")
root = tk.Tk()
root.withdraw()
root.lift()
root.attributes("-topmost", True)

nome_padrao = "leiloes_formatado.xlsx"
caminho_arquivo = filedialog.asksaveasfilename(
    initialfile=nome_padrao,
    defaultextension=".xlsx",
    filetypes=[("Planilhas Excel", "*.xlsx")],
    title="Salvar planilha como"
)

if caminho_arquivo:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leilões"

        # Estilos gerais
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        cell_font = Font(color="000000")
        cell_alignment = Alignment(wrap_text=True, vertical="top")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Lista de colunas com links para formatação de hyperlink
        link_columns = ["Link do Edital", "Link do Imóvel"] + docs_padrao

        # Escreve os dados na planilha
        for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                header_name = df.columns[c_idx - 1] if r_idx > 1 else None
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = cell_alignment

                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = cell_font
                    if header_name in link_columns and str(value).strip().lower().startswith("http"):
                        cell.hyperlink = value
                        cell.font = Font(color="0000FF", underline="single")

        # Autoajuste de largura das colunas (limite máximo de 80)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 5, 80)

        wb.save(caminho_arquivo)
        print(f"[SUCESSO] Planilha XLSX salva com sucesso em:\n{caminho_arquivo}")

    except Exception as e:
        print(f"[ERRO] Erro ao salvar a planilha: {e}")
else:
    print("[CANCELADO] Nenhum arquivo foi salvo.")

# ============================================================
# 7. Finaliza o Navegador
# ============================================================
print_header("Finalizando o Scraping e Fechando o Navegador...")
driver.quit()
